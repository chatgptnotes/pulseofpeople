"""
Excel Export Utility using openpyxl
Generates formatted Excel workbooks with multiple sheets, charts, and styling
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


class ExcelExporter:
    """Excel Report Exporter"""

    def __init__(self, report_name, data):
        self.report_name = report_name
        self.data = data
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

    def _apply_header_style(self, worksheet, row=1):
        """Apply header styling to first row"""
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_summary_sheet(self):
        """Add summary/overview sheet"""
        ws = self.wb.create_sheet("Summary")

        # Title
        ws['A1'] = self.report_name
        ws['A1'].font = Font(size=16, bold=True, color="1E3A8A")
        ws.merge_cells('A1:D1')

        # Generated date
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(size=10, color="666666")

        # Report period
        if self.data.get('date_from') and self.data.get('date_to'):
            ws['A3'] = f"Period: {self.data['date_from']} to {self.data['date_to']}"

        # Key metrics
        summary = self.data.get('summary', {})
        if summary:
            row = 5
            ws[f'A{row}'] = "Key Metrics"
            ws[f'A{row}'].font = Font(size=14, bold=True, color="1E40AF")

            row += 2
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            self._apply_header_style(ws, row)

            row += 1
            for key, value in summary.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1

        self._auto_adjust_columns(ws)

    def add_data_sheet(self, sheet_name, data, include_chart=False):
        """Add a data sheet with optional chart"""
        ws = self.wb.create_sheet(sheet_name)

        if not data or len(data) == 0:
            ws['A1'] = "No data available"
            return

        # Headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        self._apply_header_style(ws, 1)

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(data)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Alternate row colors
        gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for row_idx in range(2, len(data) + 2):
            if row_idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = gray_fill

        self._auto_adjust_columns(ws)

        # Add chart if requested
        if include_chart and len(data) > 0:
            self._add_chart_to_sheet(ws, len(data), len(headers))

    def _add_chart_to_sheet(self, worksheet, data_rows, data_cols):
        """Add chart to worksheet"""
        try:
            # Bar chart for first two columns
            chart = BarChart()
            chart.title = worksheet.title
            chart.style = 10
            chart.y_axis.title = 'Value'
            chart.x_axis.title = 'Category'

            data_ref = Reference(worksheet, min_col=2, min_row=1, max_row=min(data_rows + 1, 20), max_col=2)
            cats_ref = Reference(worksheet, min_col=1, min_row=2, max_row=min(data_rows + 1, 20))

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # Position chart
            worksheet.add_chart(chart, f"E2")
        except Exception as e:
            print(f"Error adding chart: {e}")

    def add_pivot_summary(self, sheet_name, pivot_data):
        """Add pivot table summary"""
        ws = self.wb.create_sheet(sheet_name)

        # Title
        ws['A1'] = sheet_name
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')

        row = 3
        for category, items in pivot_data.items():
            # Category header
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = Font(size=12, bold=True, color="1E40AF")
            ws.merge_cells(f'A{row}:C{row}')
            row += 1

            # Sub-items
            for key, value in items.items():
                ws[f'B{row}'] = key
                ws[f'C{row}'] = value
                row += 1

            row += 1  # Empty row between categories

        self._auto_adjust_columns(ws)

    def generate(self):
        """Generate complete Excel workbook"""
        # Summary sheet
        self.add_summary_sheet()

        # Add data sheets based on report type
        tables = self.data.get('tables', [])
        for table_info in tables:
            sheet_name = table_info.get('title', 'Data')[:31]  # Excel sheet name limit
            data = table_info.get('data', [])

            if data and len(data) > 1:
                # Convert to list of dicts if it's a list of lists
                if isinstance(data[0], list):
                    headers = data[0]
                    dict_data = []
                    for row in data[1:]:
                        dict_data.append(dict(zip(headers, row)))
                    data = dict_data

                self.add_data_sheet(sheet_name, data, include_chart=True)

        # Save to buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer


def export_voters_to_excel(voters_data):
    """Export voter data to Excel"""
    exporter = ExcelExporter("Voters Export", {})

    # Prepare data
    voter_list = []
    for voter in voters_data:
        voter_list.append({
            "Name": voter.get('name'),
            "Phone": voter.get('phone'),
            "Email": voter.get('email'),
            "Constituency": voter.get('constituency'),
            "Sentiment": voter.get('sentiment'),
            "Age": voter.get('age'),
            "Gender": voter.get('gender'),
        })

    exporter.add_data_sheet("Voters", voter_list)
    return exporter.generate()


def export_analytics_to_excel(analytics_data):
    """Export analytics data to Excel"""
    report_name = analytics_data.get('report_name', 'Analytics Report')
    exporter = ExcelExporter(report_name, analytics_data)
    return exporter.generate()


def export_custom_data_to_excel(data, sheet_configs):
    """
    Export custom data with multiple sheets
    sheet_configs: [{"name": "Sheet1", "data": [...], "include_chart": True}, ...]
    """
    exporter = ExcelExporter("Custom Export", data)

    for config in sheet_configs:
        sheet_name = config.get('name', 'Data')
        sheet_data = config.get('data', [])
        include_chart = config.get('include_chart', False)

        exporter.add_data_sheet(sheet_name, sheet_data, include_chart)

    return exporter.generate()
